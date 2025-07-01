import pandas as pd
from openpyxl import load_workbook
from colormath.color_objects import LabColor, sRGBColor, HSLColor, LCHabColor
from colormath.color_diff import delta_e_cie2000
from colormath.color_conversions import convert_color
import numpy as np

# Monkey-patch для numpy.asscalar, который был удален в новых версиях.
# colormath 3.0.0 все еще его использует.
if not hasattr(np, 'asscalar'):
    np.asscalar = lambda x: x.item()

def get_rgb_from_lab(lab_color):
    """Преобразует объект LabColor в кортеж RGB."""
    srgb = convert_color(lab_color, sRGBColor)
    # Ограничиваем значения в диапазоне 0-1 перед конвертацией
    clamped_rgb = tuple(max(0, min(1, val)) for val in srgb.get_value_tuple())
    return tuple(int(c * 255) for c in clamped_rgb)

def get_hsl_from_lab(lab_color):
    """Преобразует объект LabColor в кортеж HSL."""
    hsl = convert_color(lab_color, HSLColor)
    return hsl.get_value_tuple()

def extract_colors_from_excel(file_path):
    """
    Извлекает данные о цветах из указанного Excel-файла,
    обрабатывая ссылки на общие строки и пропуская заголовки.
    """
    all_colors = []
    try:
        # data_only=True необходимо, чтобы openpyxl разрешал ссылки (shared strings)
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        
        if 'Каталог Folio' not in workbook.sheetnames:
            print("Лист 'Каталог Folio' не найден в файле.")
            return []
            
        sheet = workbook['Каталог Folio']

        for row in sheet.iter_rows(min_row=3):
            name = row[3].value
            l_val = row[42].value
            a_val = row[44].value
            b_val = row[46].value

            if name and l_val is not None and a_val is not None and b_val is not None:
                try:
                    lab = LabColor(float(l_val), float(a_val), float(b_val))
                    all_colors.append({"name": str(name), "lab": lab})
                except (ValueError, TypeError):
                    continue
        
        print(f"Успешно извлечено {len(all_colors)} цветов из файла.")
        return all_colors

    except Exception as e:
        print(f"Произошла ошибка при чтении Excel файла: {e}")
        return []

def find_closest_color_hybrid(target_lab, color_catalog, lightness_tolerance=5.0):
    """
    Находит ближайший цвет по гибридному алгоритму:
    1. Фильтрует каталог по близкой светлоте.
    2. В отфильтрованной группе находит ближайший цвет по формуле CIEDE2000.
    """
    target_l = target_lab.lab_l
    
    # 1. Фильтрация по светлоте
    filtered_catalog = [
        color for color in color_catalog 
        if abs(color['lab'].lab_l - target_l) <= lightness_tolerance
    ]
    
    if not filtered_catalog:
        # Если в допуске ничего не найдено, ищем по всему каталогу
        filtered_catalog = color_catalog

    if not filtered_catalog:
        return None # Произойдет, только если исходный каталог пуст

    # 2. Поиск ближайшего в отфильтрованной группе
    return min(filtered_catalog, key=lambda color: delta_e_cie2000(target_lab, color['lab']))

def generate_color_scale_html(ideal_scale, real_scale, filename="color_scale.html"):
    """Генерирует HTML-файл для визуализации двух цветовых шкал."""
    
    def generate_strip(title, scale_data):
        """Вспомогательная функция для генерации одной полосы градиента."""
        strip_html = f"<h2>{title}</h2><div class='container'>"
        for item in scale_data:
            rgb = item['rgb']
            lab_l, lab_a, lab_b = item['lab'].lab_l, item['lab'].lab_a, item['lab'].lab_b
            hsl_h, hsl_s, hsl_l = item['hsl']
            text_color = "black" if lab_l > 60 else "white"
            
            info_html = f"<p>LAB: {lab_l:.1f}, {lab_a:.1f}, {lab_b:.1f}</p>"
            info_html += f"<p>HSL: {hsl_h:.0f}°, {hsl_s*100:.0f}%, {hsl_l*100:.0f}%</p>"
            
            if 'match_name' in item:
                info_html += f"<p>Folio: {item['match_name']}</p>"
                info_html += f"<p><small>dE2000: {item['delta_e']:.2f}</small></p>"

            strip_html += f"""
            <div class="color-band" style="background-color: rgb({rgb[0]}, {rgb[1]}, {rgb[2]}); color: {text_color};">
                <div class="color-info">
                    {info_html}
                </div>
            </div>"""
        strip_html += "</div>"
        return strip_html

    html_content = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>Color Scale Comparison</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; display: flex; flex-direction: column; justify-content: center; align-items: center; min-height: 100vh; background-color: #f4f4f4; margin: 1em 0; padding: 1em; }
        .container { display: flex; flex-direction: row; border: 1px solid #ddd; box-shadow: 0 4px 12px rgba(0,0,0,0.1); margin-bottom: 2em; border-radius: 8px; overflow: hidden; }
        .color-band { padding: 12px; min-height: 120px; width: 110px; display: flex; align-items: center; justify-content: center; text-align: center; }
        .color-info { text-shadow: 0 1px 2px rgba(0,0,0,0.4); }
        .color-info p { margin: 4px 0; font-size: 0.8rem; }
        .color-info small { font-size: 0.7rem; opacity: 0.8; }
        h2 { text-align: center; font-weight: 500; color: #333; }
    </style>
</head>
<body>
"""
    html_content += generate_strip("Ideal Gradient (Interpolated)", ideal_scale)
    html_content += generate_strip("Real Folio Colors (Closest Match)", real_scale)

    html_content += """
</body>
</html>"""
    with open(filename, "w", encoding='utf-8') as f:
        f.write(html_content)


def main():
    """
    Главная функция для генерации градиентной шкалы по методу попеременных шагов.
    """
    excel_file = '27.06.2025г. Каталог Folio (составы) .xlsx'
    
    color_catalog = extract_colors_from_excel(excel_file)
    
    if not color_catalog:
        print("Не удалось извлечь цвета из каталога. Завершение работы.")
        return

    # Крайние точки из задания
    start_lab = LabColor(22.260, 3.294, -5.936)
    end_lab = LabColor(92.5239, 1.0497, -1.8174)
    
    # Конвертируем в LCH для работы с тоном, насыщенностью и светлотой
    start_lch = convert_color(start_lab, LCHabColor)
    end_lch = convert_color(end_lab, LCHabColor)

    # Рассчитываем средний тон (с учетом "перехода через 0/360 градусов")
    h1, h2 = start_lch.lch_h, end_lch.lch_h
    avg_hue = (h1 + h2) / 2
    if abs(h1 - h2) > 180:
        avg_hue = (avg_hue + 180) % 360

    num_points = 11  # 10 шагов = 11 точек в градиенте
    # Распределяем количество шагов изменения между светлотой и насыщенностью
    lightness_steps_count = (num_points - 1) // 2 + ((num_points - 1) % 2)
    chroma_steps_count = (num_points - 1) // 2
    
    delta_l = end_lch.lch_l - start_lch.lch_l
    delta_c = end_lch.lch_c - start_lch.lch_c

    l_step_size = delta_l / lightness_steps_count if lightness_steps_count > 0 else 0
    c_step_size = delta_c / chroma_steps_count if chroma_steps_count > 0 else 0

    ideal_color_scale = []
    real_color_scale = []

    l_val, c_val = start_lch.lch_l, start_lch.lch_c

    for i in range(num_points):
        if i == 0:
            gen_lch = start_lch
        elif i == num_points - 1:
            # Гарантируем точное попадание в конечный цвет
            gen_lch = end_lch
        else:
            # Попеременно меняем L и C
            if (i - 1) % 2 == 0:  # Шаги 1, 3, 5... (переходы 0, 2, 4...) -> меняем L
                l_val += l_step_size
            else:  # Шаги 2, 4, 6... (переходы 1, 3, 5...) -> меняем C
                c_val += c_step_size
            gen_lch = LCHabColor(l_val, c_val, avg_hue)
        
        gen_lab = convert_color(gen_lch, LabColor)

        # Добавляем теоретический цвет в шкалу
        ideal_rgb = get_rgb_from_lab(gen_lab)
        ideal_hsl = get_hsl_from_lab(gen_lab)
        ideal_color_scale.append({
            "step": i + 1,
            "lab": gen_lab,
            "rgb": ideal_rgb,
            "hsl": ideal_hsl
        })

        # Ищем ближайший реальный цвет по гибридному методу
        closest_match = find_closest_color_hybrid(gen_lab, color_catalog, lightness_tolerance=5.0)
        
        if closest_match:
            closest_rgb = get_rgb_from_lab(closest_match['lab'])
            closest_hsl = get_hsl_from_lab(closest_match['lab'])
            delta_e = delta_e_cie2000(gen_lab, closest_match['lab'])
            real_color_scale.append({
                "step": i + 1,
                "lab": closest_match['lab'],
                "rgb": closest_rgb,
                "hsl": closest_hsl,
                "match_name": closest_match['name'],
                "delta_e": delta_e
            })

    generate_color_scale_html(ideal_color_scale, real_color_scale, "index.html")
    print("Генерация index.html завершена.")

if __name__ == "__main__":
    main() 